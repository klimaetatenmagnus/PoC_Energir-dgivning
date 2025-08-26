import openpyxl
from typing import List, Dict, Optional

class StotteordningFinner:
    def __init__(self, excel_path: str):
        self.wb = openpyxl.load_workbook(excel_path, data_only=True)
        self.sheet = self.wb.active
        
        # Definer tiltak og deres kolonneposisjoner
        self.tiltak_kolonner = {
            "tetting": {"start": 5, "bygningstyper": ["enebolig", "rekkehus", "blokk"]},
            "solenergi": {"start": 8, "bygningstyper": ["enebolig", "rekkehus", "blokk"]},
            "etterisolering_fasade": {"start": 11, "bygningstyper": ["enebolig", "rekkehus", "blokk"]},
            "etterisolering_kjeller_loft": {"start": 14, "bygningstyper": ["enebolig", "rekkehus", "blokk"]},
            "varmepumpe": {"start": 17, "bygningstyper": ["enebolig", "rekkehus", "blokk"]},
            "ventilasjon": {"start": 20, "bygningstyper": ["enebolig", "rekkehus", "blokk"]},
            "vinduer": {"start": 23, "bygningstyper": ["enebolig", "rekkehus", "blokk"]},
            "smart_energistyring": {"start": 26, "bygningstyper": ["enebolig", "rekkehus", "blokk"]},
            "annen_oppvarming": {"start": 29, "bygningstyper": ["enebolig", "rekkehus", "blokk"]}
        }
        
        # Definer bygningstype kolonneindeks
        self.bygningstype_offset = {
            "enebolig": 0,
            "rekkehus": 1,
            "blokk": 2
        }
        
        # Start- og sluttrad for gulliste og ikke-gulliste
        self.gulliste_start = 7
        self.gulliste_slutt = 50
        self.ikke_gulliste_start = 52
        self.ikke_gulliste_slutt = 80  # Antar maks 150 rader
        
    def finn_stotteordninger(self, gulliste: bool, tiltak: str, bygningstype: str) -> List[Dict[str, str]]:
        """
        Finner støtteordninger basert på gulliste-status, tiltak og bygningstype.
        
        Args:
            gulliste: True hvis bygget er på gul liste, False ellers
            tiltak: Type tiltak (f.eks. "solenergi", "tetting", "varmepumpe")
            bygningstype: Type bygg ("enebolig", "rekkehus" eller "blokk")
            
        Returns:
            Liste med dictionaries som inneholder ordningsnavn og lenke
        """
        resultater = []
        
        # Valider input
        if tiltak not in self.tiltak_kolonner:
            raise ValueError(f"Ukjent tiltak: {tiltak}. Gyldige tiltak: {list(self.tiltak_kolonner.keys())}")
        
        if bygningstype not in self.bygningstype_offset:
            raise ValueError(f"Ukjent bygningstype: {bygningstype}. Gyldige typer: {list(self.bygningstype_offset.keys())}")
        
        # Finn riktig kolonne
        tiltak_info = self.tiltak_kolonner[tiltak]
        kolonne = tiltak_info["start"] + self.bygningstype_offset[bygningstype]
        
        # Bestem radområde basert på gulliste
        if gulliste:
            start_rad = self.gulliste_start
            slutt_rad = self.gulliste_slutt
        else:
            start_rad = self.ikke_gulliste_start
            slutt_rad = self.ikke_gulliste_slutt
        
        # Hold styr på gjeldende overskrift
        overskrifter = ["Klima- og energifondet", "Enova", "Byantikvaren", "Riksantikvaren", "Kulturminnefondet", "Klimaetaten"]
        gjeldende_overskrift = None
        
        # Gå gjennom rader og finn markerte celler
        for rad in range(start_rad, slutt_rad + 1):
            # Sjekk om denne raden har en overskrift
            ordning_celle = self.sheet.cell(row=rad, column=3)  # Kolonne C
            if ordning_celle.value:
                celle_verdi = str(ordning_celle.value).strip()
                
                # Sjekk om dette er en ren overskrift (eksakt match eller nesten eksakt)
                if celle_verdi == "Klima- og energifondet":
                    gjeldende_overskrift = "Klima- og energifondet"
                elif celle_verdi == "Enova":
                    gjeldende_overskrift = "Enova"
                elif celle_verdi == "Byantikvaren":
                    gjeldende_overskrift = "Byantikvaren"
                elif celle_verdi == "Riksantikvaren":
                    gjeldende_overskrift = "Riksantikvaren"
                elif celle_verdi == "Kulturminnefondet":
                    gjeldende_overskrift = "Kulturminnefondet"
                # Hvis det ikke er en ren overskrift, sjekk om det er en støtteordning med "Enova" i navnet
                elif "| Enova" in celle_verdi:
                    # Dette er en Enova-støtteordning, ikke en overskriftsendring
                    pass
                # Sjekk om det er en Klimaetaten-støtteordning
                elif "- Klimaetaten" in celle_verdi and gjeldende_overskrift != "Enova":
                    # Hvis vi er under Klima- og energifondet, behold den overskriften
                    if gjeldende_overskrift != "Klima- og energifondet":
                        gjeldende_overskrift = "Klima- og energifondet"
            
            # Sjekk om cellen er markert (inneholder X eller annen verdi)
            celle = self.sheet.cell(row=rad, column=kolonne)
            if celle.value and str(celle.value).strip() not in ["", "None"]:
                # Finn ordningsnavn og lenke
                belop_celle = self.sheet.cell(row=rad, column=4)   # Kolonne D har beløp/info
                
                if ordning_celle.value:
                    ordning_info = {
                        "ordning": str(ordning_celle.value),
                        "lenke": None,
                        "belop": str(belop_celle.value) if belop_celle.value else None,
                        "rad": rad,
                        "overskrift": gjeldende_overskrift
                    }
                    
                    # Sjekk om det er en hyperlink
                    if ordning_celle.hyperlink:
                        ordning_info["lenke"] = ordning_celle.hyperlink.target
                    
                    resultater.append(ordning_info)
        
        return resultater
    
    def print_resultater(self, resultater: List[Dict[str, str]], gulliste: bool, tiltak: str, bygningstype: str):
        """Printer resultatene på en formatert måte"""
        print(f"\n{'='*80}")
        print(f"Støtteordninger for:")
        print(f"  - Gulliste: {'Ja' if gulliste else 'Nei'}")
        print(f"  - Tiltak: {tiltak}")
        print(f"  - Bygningstype: {bygningstype}")
        print(f"{'='*80}")
        
        if not resultater:
            print("Ingen støtteordninger funnet for denne kombinasjonen.")
        else:
            for i, ordning in enumerate(resultater, 1):
                print(f"\n{i}. {ordning['ordning']}")
                if ordning.get('overskrift'):
                    print(f"   Under: {ordning['overskrift']}")
                if ordning['belop']:
                    print(f"   Beløp/info: {ordning['belop']}")
                if ordning['lenke']:
                    print(f"   Lenke: {ordning['lenke']}")
                print(f"   (Fra rad {ordning['rad']})")


def main():
    # Opprett objekt
    finder = StotteordningFinner("(2) Matrise_Energitiltak og relevante støtteordninger.xlsx")
    
    # Test case 1: Gulliste=True, Solenergi, Rekkehus
    print("\n\nTEST CASE 1:")
    resultater1 = finder.finn_stotteordninger(
        gulliste=True,
        tiltak="solenergi",
        bygningstype="rekkehus"
    )
    finder.print_resultater(resultater1, True, "solenergi", "rekkehus")
    
    # Test case 2: Gulliste=False, Solenergi, Enebolig
    print("\n\nTEST CASE 2:")
    resultater2 = finder.finn_stotteordninger(
        gulliste=False,
        tiltak="solenergi",
        bygningstype="enebolig"
    )
    finder.print_resultater(resultater2, False, "solenergi", "enebolig")
    
    # Test case 3: Gulliste=True, Tetting, Blokk
    print("\n\nTEST CASE 3:")
    resultater3 = finder.finn_stotteordninger(
        gulliste=True,
        tiltak="tetting",
        bygningstype="blokk"
    )
    finder.print_resultater(resultater3, True, "tetting", "blokk")


if __name__ == "__main__":
    main()