import openpyxl
import json
from typing import Dict, List, Any
import os
from datetime import datetime

class StotteordningCache:
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.cache_file = "stotteordninger_data.json"
        self.wb = None
        self.sheet = None
        self.data = {}
        
        # Definer tiltak og deres kolonneposisjoner
        self.tiltak_kolonner = {
            "tetting": {"start": 5, "bygningstyper": ["enebolig", "rekkehus", "blokk"]},
            "solenergi": {"start": 8, "bygningstyper": ["enebolig", "rekkehus", "blokk"]},
            "etterisolering_fasade": {"start": 11, "bygningstyper": ["enebolig", "rekkehus", "blokk"]},
            "etterisolering_kjeller_loft": {"start": 14, "bygningstyper": ["enebolig", "rekkehus", "blokk"]},
            "varmepumpe": {"start": 17, "bygningstyper": ["enebolig", "rekkehus", "blokk"]},
            "ventilasjon": {"start": 20, "bygningstyper": ["enebolig", "rekkehus", "blokk"]},
            "vinduer": {"start": 23, "bygningstyper": ["enebolig", "rekkehus", "blokk"]},
            "smart_energistyring": {"start": 26, "bygningstyper": ["enebolig", "rekkehus", "blokk"]},
            "annen_oppvarming": {"start": 29, "bygningstyper": ["enebolig", "rekkehus", "blokk"]}
        }
        
        # Definer bygningstype kolonneindeks
        self.bygningstype_offset = {
            "enebolig": 0,
            "rekkehus": 1,
            "blokk": 2
        }
        
        # Start- og sluttrad for gulliste og ikke-gulliste
        self.gulliste_start = 7
        self.gulliste_slutt = 50
        self.ikke_gulliste_start = 52
        self.ikke_gulliste_slutt = 80
        
    def oppdater_cache(self):
        """Oppdaterer cache fra Excel-filen"""
        print("Oppdaterer cache fra Excel...")
        self.wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        self.sheet = self.wb.active
        
        # Bygg opp dictionary struktur
        self.data = {
            "sist_oppdatert": datetime.now().isoformat(),
            "gulliste": self._hent_alle_stotteordninger(True),
            "ikke_gulliste": self._hent_alle_stotteordninger(False)
        }
        
        # Lagre til JSON-fil
        with open(self.cache_file, 'w', encoding='utf-8') as f:
            json.dump(self.data, f, ensure_ascii=False, indent=2)
        
        print(f"Cache oppdatert og lagret til {self.cache_file}")
        return self.data
    
    def _hent_alle_stotteordninger(self, gulliste: bool) -> Dict[str, Dict[str, List[Dict[str, Any]]]]:
        """Henter alle støtteordninger for gulliste/ikke-gulliste"""
        resultat = {}
        
        # Bestem radområde basert på gulliste
        if gulliste:
            start_rad = self.gulliste_start
            slutt_rad = self.gulliste_slutt
        else:
            start_rad = self.ikke_gulliste_start
            slutt_rad = self.ikke_gulliste_slutt
        
        # Gå gjennom alle tiltak
        for tiltak, tiltak_info in self.tiltak_kolonner.items():
            resultat[tiltak] = {}
            
            # Gå gjennom alle bygningstyper
            for bygningstype in self.bygningstype_offset.keys():
                kolonne = tiltak_info["start"] + self.bygningstype_offset[bygningstype]
                stotteordninger = []
                
                # Hold styr på gjeldende overskrift
                overskrifter = ["Klima- og energifondet", "Enova", "Byantikvaren", "Riksantikvaren", "Kulturminnefondet", "Klimaetaten"]
                gjeldende_overskrift = None
                
                # Gå gjennom rader og finn markerte celler
                for rad in range(start_rad, slutt_rad + 1):
                    # Sjekk om denne raden har en overskrift
                    ordning_celle = self.sheet.cell(row=rad, column=3)  # Kolonne C
                    if ordning_celle.value:
                        celle_verdi = str(ordning_celle.value).strip()
                        
                        # Sjekk om dette er en ren overskrift (eksakt match eller nesten eksakt)
                        if celle_verdi == "Klima- og energifondet":
                            gjeldende_overskrift = "Klima- og energifondet"
                        elif celle_verdi == "Enova":
                            gjeldende_overskrift = "Enova"
                        elif celle_verdi == "Byantikvaren":
                            gjeldende_overskrift = "Byantikvaren"
                        elif celle_verdi == "Riksantikvaren":
                            gjeldende_overskrift = "Riksantikvaren"
                        elif celle_verdi == "Kulturminnefondet":
                            gjeldende_overskrift = "Kulturminnefondet"
                        # Hvis det ikke er en ren overskrift, sjekk om det er en støtteordning med "Enova" i navnet
                        elif "| Enova" in celle_verdi:
                            # Dette er en Enova-støtteordning, ikke en overskriftsendring
                            pass
                        # Sjekk om det er en Klimaetaten-støtteordning
                        elif "- Klimaetaten" in celle_verdi and gjeldende_overskrift != "Enova":
                            # Hvis vi er under Klima- og energifondet, behold den overskriften
                            if gjeldende_overskrift != "Klima- og energifondet":
                                gjeldende_overskrift = "Klima- og energifondet"
                    
                    # Sjekk om cellen er markert
                    celle = self.sheet.cell(row=rad, column=kolonne)
                    if celle.value and str(celle.value).strip() not in ["", "None"]:
                        # Finn ordningsnavn og lenke
                        belop_celle = self.sheet.cell(row=rad, column=4)  # Kolonne D
                        
                        if ordning_celle.value:
                            ordning_info = {
                                "ordning": str(ordning_celle.value),
                                "lenke": None,
                                "belop": str(belop_celle.value) if belop_celle.value else None,
                                "overskrift": gjeldende_overskrift
                            }
                            
                            # Sjekk om det er en hyperlink
                            if ordning_celle.hyperlink:
                                ordning_info["lenke"] = ordning_celle.hyperlink.target
                            
                            stotteordninger.append(ordning_info)
                
                resultat[tiltak][bygningstype] = stotteordninger
        
        return resultat
    
    def hent_stotteordninger(self, gulliste: bool, tiltak: str, bygningstype: str) -> List[Dict[str, Any]]:
        """Henter støtteordninger fra cache"""
        # Last cache fra fil hvis den finnes
        if os.path.exists(self.cache_file):
            with open(self.cache_file, 'r', encoding='utf-8') as f:
                self.data = json.load(f)
        else:
            # Opprett cache hvis den ikke finnes
            self.oppdater_cache()
        
        # Hent fra cache
        gulliste_key = "gulliste" if gulliste else "ikke_gulliste"
        
        try:
            return self.data[gulliste_key][tiltak][bygningstype]
        except KeyError:
            return []
    
    def eksporter_til_javascript(self):
        """Eksporterer data til en JavaScript-fil som kan brukes direkte i React"""
        if not self.data:
            self.oppdater_cache()
        
        js_content = f"""// Auto-generert fil fra stotteordning_cache.py
// Sist oppdatert: {self.data['sist_oppdatert']}

export const stotteordningData = {json.dumps(self.data, ensure_ascii=False, indent=2)};

export function getStotteordninger(gulliste, tiltak, bygningstype) {{
  const gullisteKey = gulliste ? 'gulliste' : 'ikke_gulliste';
  
  try {{
    return stotteordningData[gullisteKey][tiltak][bygningstype] || [];
  }} catch (error) {{
    console.error('Kunne ikke hente støtteordninger:', error);
    return [];
  }}
}}
"""
        
        # Generer filen direkte i src/data/ mappen
        js_file = "src/data/stotteordningData.js"
        os.makedirs(os.path.dirname(js_file), exist_ok=True)
        with open(js_file, 'w', encoding='utf-8') as f:
            f.write(js_content)
        
        print(f"JavaScript-fil generert: {js_file}")
        return js_file


def main():
    # Test cache-funksjonen
    cache = StotteordningCache("(2) Matrise_Energitiltak og relevante støtteordninger.xlsx")
    
    # Oppdater cache
    cache.oppdater_cache()
    
    # Eksporter til JavaScript
    cache.eksporter_til_javascript()
    
    # Test henting fra cache
    print("\n\nTest henting fra cache:")
    resultater = cache.hent_stotteordninger(
        gulliste=True,
        tiltak="solenergi",
        bygningstype="enebolig"
    )
    
    print(f"\nFunnet {len(resultater)} støtteordninger for gulliste=True, solenergi, enebolig:")
    for ordning in resultater:
        print(f"- {ordning['ordning']}")
        if ordning.get('overskrift'):
            print(f"  Under: {ordning['overskrift']}")


if __name__ == "__main__":
    main()