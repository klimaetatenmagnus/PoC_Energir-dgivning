import openpyxl

# Åpne Excel-filen
wb = openpyxl.load_workbook("(2) Matrise_Energitiltak og relevante støtteordninger.xlsx")
sheet = wb.active

print("=== ANALYSE AV EXCEL-STRUKTUR ===\n")

# Finn overskrifter og seksjonsskiller
print("RAD-INNHOLD (første 100 rader):")
print("-" * 80)

gulliste_start = None
gulliste_slutt = None
ikke_gulliste_start = None
ikke_gulliste_slutt = None
overskrifter = []

for row in range(1, 101):
    # Sjekk kolonne A, B og C for innhold
    col_a = sheet.cell(row=row, column=1).value
    col_b = sheet.cell(row=row, column=2).value
    col_c = sheet.cell(row=row, column=3).value
    
    # Se etter viktige markører
    if col_a or col_b or col_c:
        content = f"Rad {row}: "
        if col_a:
            content += f"A: {col_a} | "
        if col_b:
            content += f"B: {col_b} | "
        if col_c:
            content += f"C: {col_c}"
        print(content)
        
        # Se etter "Gul liste" markør
        if any(val and "gul liste" in str(val).lower() for val in [col_a, col_b, col_c]):
            if not gulliste_start:
                print(f"  --> GULLISTE START FUNNET!")
                gulliste_start = row + 1  # Start på neste rad
        
        # Se etter "Ikke gul liste" eller lignende markør
        if any(val and ("ikke" in str(val).lower() and "gul" in str(val).lower()) for val in [col_a, col_b, col_c]):
            if gulliste_start and not gulliste_slutt:
                gulliste_slutt = row - 1
                print(f"  --> GULLISTE SLUTT: {gulliste_slutt}")
            ikke_gulliste_start = row + 1
            print(f"  --> IKKE-GULLISTE START FUNNET!")
        
        # Identifiser mulige overskrifter (Enova, Klima- og energifondet, etc.)
        if col_c and any(overskrift in str(col_c) for overskrift in ["Enova", "Klima", "Byantikvaren", "Riksantikvaren", "Kulturminnefondet"]):
            overskrifter.append({"rad": row, "tekst": str(col_c)})
            print(f"  --> OVERSKRIFT FUNNET: {col_c}")

# Finn siste rad med innhold
last_content_row = 1
for row in range(100, 1, -1):
    has_content = False
    for col in range(1, 32):  # Sjekk kolonner A-AE
        if sheet.cell(row=row, column=col).value:
            has_content = True
            break
    if has_content:
        last_content_row = row
        break

if ikke_gulliste_start and not ikke_gulliste_slutt:
    ikke_gulliste_slutt = last_content_row

print("\n" + "=" * 80)
print("\nOPPSUMMERING:")
print(f"Gulliste rader: {gulliste_start} - {gulliste_slutt}")
print(f"Ikke-gulliste rader: {ikke_gulliste_start} - {ikke_gulliste_slutt}")
print(f"\nFunnet overskrifter:")
for overskrift in overskrifter:
    print(f"  Rad {overskrift['rad']}: {overskrift['tekst']}")

# Analyser kolonnestruktur
print("\n" + "=" * 80)
print("\nKOLONNESTRUKTUR (rad 5-7):")
for row in range(5, 8):
    print(f"\nRad {row}:")
    for col in range(4, 32):  # Kolonner D-AE
        value = sheet.cell(row=row, column=col).value
        if value:
            print(f"  Kolonne {chr(64 + col)}: {value}")

wb.close()